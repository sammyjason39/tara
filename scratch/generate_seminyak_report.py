import pandas as pd
import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

excel_catalog_path = r"C:\Users\user\Downloads\Bambu Silver\excel_catalog.json"
csv_path = r"C:\Users\user\Downloads\Bambu Silver\recovered_seminyak_opname.csv"
image_transcription_path = r"scratch\image_transcription.json"
db_images_map_path = r"scratch\item_images_map.json"
output_report_path = r"C:\Users\user\Downloads\Bambu Silver\seminyak_stock_opname_report_v2.xlsx"

print("Generating Seminyak Stock Opname Report...")
print("Note: Mapping Seminyak branch to 'UBUD 1' Excel branch.")

# 1. Load Excel Catalog
with open(excel_catalog_path, 'r', encoding='utf-8') as f:
    excel_catalog = json.load(f)

# 2. Parse CSV Scans
csv_counts = {}
if os.path.exists(csv_path):
    with open(csv_path, 'r', encoding='utf-8') as f:
        # Skip header
        lines = f.readlines()[1:]
        for line in lines:
            line = line.strip()
            if not line:
                continue
            # Handle quoted CSV fields
            match = re.match(r'^"([^"]*)","([^"]*)",(\d+)', line)
            if match:
                sku = match.group(1).strip()
                actual = int(match.group(3))
                if sku:
                    csv_counts[sku] = csv_counts.get(sku, 0) + actual
else:
    print(f"Warning: CSV scans not found at {csv_path}")

# 3. Load Image Counts
image_counts = {}
if os.path.exists(image_transcription_path):
    with open(image_transcription_path, 'r', encoding='utf-8') as f:
        image_counts = json.load(f)
else:
    print(f"Warning: Image transcription JSON not found at {image_transcription_path}")

# Load DB Images Map
db_images_map = {}
if os.path.exists(db_images_map_path):
    with open(db_images_map_path, 'r', encoding='utf-8') as f:
        db_images_map = json.load(f)

# Helper to normalize SKU for fallback matching
def normalize_sku(sku):
    if not sku:
        return ""
    return sku.strip().upper().replace('.', '').replace('-', '').replace(' ', '')

db_normalized = {}
for db_sku, info in db_images_map.items():
    norm = normalize_sku(db_sku)
    if norm:
        db_normalized[norm] = info

# 4. Merge Opname Counts
opname_counts = {**csv_counts}
for sku, count in image_counts.items():
    opname_counts[sku] = opname_counts.get(sku, 0) + count

print(f"Unique SKUs counted: {len(opname_counts)}")

# 5. Index Excel catalog
excel_by_sku = {}
excel_seminyak_by_sku = {}

for record in excel_catalog:
    sku = record['sku']
    branch = record['branch']
    
    # Store in master catalog by SKU
    # Prefer records with pricing if duplicates exist
    if sku not in excel_by_sku or (excel_by_sku[sku]['capital'] == 0 and record['capital'] > 0):
        excel_by_sku[sku] = record
        
    # Store Seminyak-specific records (using 'UBUD 1' branch)
    if branch == 'UBUD 1':
        excel_seminyak_by_sku[sku] = record

print(f"Master catalog unique items: {len(excel_by_sku)}")
print(f"Seminyak branch ('UBUD 1') items in Excel: {len(excel_seminyak_by_sku)}")

# Manual/New Items Definitions
manual_items = {
    '585 557R': {
        'name': 'E.ZIRCONE MIX STAR 6MM (RED)',
        'category': 'ZIRCONIA',
        'capital': 21550,
        'selling': 200000
    },
    '580 209C': {
        'name': 'P.OPAL STONE W/RESIN C',
        'category': 'STONES',
        'capital': 244750,
        'selling': 890000
    }
}

# 6. Build Report Dataset
all_skus = set(excel_seminyak_by_sku.keys()).union(opname_counts.keys())
rows = []

for sku in all_skus:
    # Resolve catalog info
    cat_info = excel_by_sku.get(sku)
    
    # Fallback to manual registration
    if not cat_info and sku in manual_items:
        cat_info = manual_items[sku]
        
    name = cat_info['name'] if cat_info else f"[Unregistered] Barcode: {sku}"
    category = cat_info['category'] if cat_info else "Uncategorized"
    capital_price = cat_info['capital'] if cat_info else 0
    selling_price = cat_info['selling'] if cat_info else 0
    
    # Get Old Quantity (from Excel branch UBUD 1)
    old_qty = excel_seminyak_by_sku[sku]['qty'] if sku in excel_seminyak_by_sku else 0
    
    # Get New/Current Quantity
    if sku in opname_counts:
        current_qty = opname_counts[sku]
    else:
        current_qty = old_qty
        
    total_capital = current_qty * capital_price
    total_selling = current_qty * selling_price
    
    # Have Picture
    has_db_pic = False
    if sku in db_images_map:
        has_db_pic = db_images_map[sku]['has_picture']
    else:
        sku_norm = normalize_sku(sku)
        if sku_norm in db_normalized:
            has_db_pic = db_normalized[sku_norm]['has_picture']
    
    have_picture = "Yes" if (has_db_pic or sku in image_counts) else "No"

    rows.append({
        'SKU': sku,
        'Name': name,
        'Category': category,
        'Old Qty': int(old_qty),
        'Current Qty': int(current_qty),
        'Capital Price': capital_price,
        'Selling Price': selling_price,
        'Total Capital': total_capital,
        'Total Selling': total_selling,
        'Have Picture': have_picture
    })

# Sort by SKU
rows = sorted(rows, key=lambda x: x['SKU'])
df_report = pd.DataFrame(rows)

print(f"Report total rows: {len(df_report)}")

# 7. Write and Style with Openpyxl
wb = Workbook()
ws = wb.active
ws.title = "Stock Opname Seminyak"
ws.views.sheetView[0].showGridLines = True

# Colors (Sleek Indigo/Slate theme)
primary_color = "3F51B5" # Indigo
header_text_color = "FFFFFF"
zebra_color = "F3F4F9"
total_fill_color = "E8EAF6"
accent_border_color = "D0D4DF"

font_title = Font(name="Segoe UI", size=16, bold=True, color="1A237E")
font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="555555")
font_header = Font(name="Segoe UI", size=11, bold=True, color=header_text_color)
font_data = Font(name="Segoe UI", size=10)
font_total = Font(name="Segoe UI", size=11, bold=True, color="1A237E")

fill_header = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
fill_zebra = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid")
fill_total = PatternFill(start_color=total_fill_color, end_color=total_fill_color, fill_type="solid")

thin_side = Side(border_style="thin", color=accent_border_color)
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side(border_style="medium", color="000000"))
border_total = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")

# Title Block
ws.merge_cells("A1:J1")
ws["A1"] = "BAMBU SILVER - SEMINYAK STOCK OPNAME REPORT"
ws["A1"].font = font_title
ws["A1"].alignment = align_left

ws.merge_cells("A2:J2")
ws["A2"] = "Location: Seminyak Branch (Ubud 1 in Excel)  |  Audit Date: May 2026"
ws["A2"].font = font_subtitle
ws["A2"].alignment = align_left

ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 10 # empty spacer row

# Headers
headers = [
    "SKU", "Item Name", "Category", 
    "Old Qty", "Current Qty", 
    "Capital Price", "Selling Price", 
    "Total Capital", "Total Selling", "Have Picture"
]

header_row = 4
ws.row_dimensions[header_row].height = 25

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=col_idx)
    cell.value = h
    cell.font = font_header
    cell.fill = fill_header
    cell.border = border_header
    if h in ["SKU", "Category", "Have Picture"]:
        cell.alignment = align_center
    elif h in ["Old Qty", "Current Qty", "Capital Price", "Selling Price", "Total Capital", "Total Selling"]:
        cell.alignment = align_right
    else:
        cell.alignment = align_left

# Data Rows
start_row = 5
for idx, r in enumerate(rows):
    current_row = start_row + idx
    ws.row_dimensions[current_row].height = 20
    is_even = idx % 2 == 1
    
    # Values
    values = [
        r['SKU'], r['Name'], r['Category'],
        r['Old Qty'], r['Current Qty'],
        r['Capital Price'], r['Selling Price'],
        r['Total Capital'], r['Total Selling'],
        r['Have Picture']
    ]
    
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = val
        cell.font = font_data
        cell.border = border_cell
        
        if is_even:
            cell.fill = fill_zebra
            
        # Alignments & Number Formats
        if col_idx in [1]: # SKU
            cell.alignment = align_center
            cell.number_format = "@"
        elif col_idx in [2]: # Name
            cell.alignment = align_left
        elif col_idx in [3]: # Category
            cell.alignment = align_center
        elif col_idx in [4, 5]: # Quantities
            cell.alignment = align_right
            cell.number_format = "#,##0"
        elif col_idx in [6, 7, 8, 9]: # Prices & Totals
            cell.alignment = align_right
            cell.number_format = "Rp #,##0"
        elif col_idx in [10]: # Have Picture
            cell.alignment = align_center

# Grand Totals Row
total_row = start_row + len(rows)
ws.row_dimensions[total_row].height = 25

# Merged Total label
ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
total_label_cell = ws.cell(row=total_row, column=1)
total_label_cell.value = "GRAND TOTAL"
total_label_cell.font = font_total
total_label_cell.alignment = Alignment(horizontal="center", vertical="center")

# Fill merged cells
for col_idx in range(1, 4):
    cell = ws.cell(row=total_row, column=col_idx)
    cell.fill = fill_total
    cell.border = border_total

# Formulas for Totals
formulas = {
    4: f"=SUM(D{start_row}:D{total_row-1})", # Old Qty
    5: f"=SUM(E{start_row}:E{total_row-1})", # Current Qty
    8: f"=SUM(H{start_row}:H{total_row-1})", # Total Capital
    9: f"=SUM(I{start_row}:I{total_row-1})"  # Total Selling
}

for col_idx in range(4, 11):
    cell = ws.cell(row=total_row, column=col_idx)
    cell.fill = fill_total
    cell.border = border_total
    cell.font = font_total
    
    if col_idx in formulas:
        cell.value = formulas[col_idx]
        
    if col_idx in [4, 5]:
        cell.alignment = align_right
        cell.number_format = "#,##0"
    elif col_idx in [8, 9]:
        cell.alignment = align_right
        cell.number_format = "Rp #,##0"
    else:
        # Price columns (6 & 7) shouldn't be summed, leave blank but styled
        cell.value = ""

# Auto-adjust column widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        # Avoid including merged cells in width calculation
        if cell.row in [1, 2, total_row] and col_letter in ['A', 'B', 'C']:
            continue
        try:
            if cell.value:
                # Add extra space for currency formatting
                val_str = str(cell.value)
                if isinstance(cell.value, (int, float)) and col_letter in ['F', 'G', 'H', 'I']:
                    val_str = f"Rp {cell.value:,.0f}"
                max_len = max(max_len, len(val_str))
        except:
            pass
    # Set padding
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save(output_report_path)
print(f"Report successfully saved to {output_report_path}!")
